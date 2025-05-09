import customtkinter as ctk
from tkinter import messagebox, ttk, Toplevel, Text, Scrollbar
from datetime import datetime
import os
import logging
from openpyxl import Workbook, load_workbook
import json
import win32print
import win32api
from win32printing import Printer
import codecs
import threading

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Constants
CONFIG_FILE = "app_config.json"
INVOICE_SAVE_DIR = r"D:\invoices"  # Default directory for saving invoices
AUTOSAVE_INTERVAL = 300000  # 5 minutes in milliseconds

# Item list for dropdown
ITEM_LIST = [
    "MAIZE", "SOYABEAN", "LOBHA", "HULLI", "KADLI", "BLACK MOONG", 
    "CHAMAKI MOONG", "RAGI", "WHEAT", "RICE", "BILAJOLA", "BIJAPUR", 
    "CHS-5", "FEEDS", "KUSUBI", "SASAVI", "SAVI", "CASTER SEEDS", 
    "TOOR RED", "TOOR WHITE", "HUNASIBIKA", "SF", "AWARI",
    "Add New Item..."  # Add this as the last option
]

def validate_float(value):
    """Validate if a string can be converted to float."""
    try:
        return float(value) if value.strip() else 0
    except ValueError:
        return 0

# Define font configurations
HEADER_FONT = ("Segoe UI", 28, "bold")
SUBHEADER_FONT = ("Segoe UI", 16)
LABEL_FONT = ("Segoe UI", 13)
ENTRY_FONT = ("Segoe UI", 13)
TABLE_HEADER_FONT = ("Segoe UI", 13, "bold")
TABLE_FONT = ("Segoe UI", 13)
BUTTON_FONT = ("Segoe UI", 13)

# Define color scheme for light theme
PRIMARY_COLOR = "#1976d2"      # Blue
SECONDARY_COLOR = "#2196f3"    # Lighter blue
ACCENT_COLOR = "#64b5f6"       # Even lighter blue
BACKGROUND_COLOR = "#ffffff"    # White
FRAME_COLOR = "#f5f5f5"        # Light gray
BORDER_COLOR = "#e0e0e0"       # Border gray
TEXT_COLOR = "#212121"         # Dark gray for text
ERROR_COLOR = "#f44336"        # Red

class InvoiceApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        # Initialize mode-specific data storage first with empty data structures
        self.mode_data = {
            "Patti": [],
            "Kata": [],
            "Barthe": []
        }
        
        # Track whether data has been entered in each mode
        self.mode_initialized = {
            "Patti": False,
            "Kata": False, 
            "Barthe": False
        }
        
        self.load_config()
        self.setup_ui()
        self._update_timer = None  # Add this line for debouncing

    def load_config(self):
        """Load application configuration from file"""
        self.config = {
            "theme": "Green",
            "window_size": "1200x800",
            "autosave": True
        }
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, 'r') as f:
                    self.config.update(json.load(f))
        except Exception as e:
            logging.error(f"Error loading config: {e}")

    def save_config(self):
        """Save application configuration to file"""
        try:
            with open(CONFIG_FILE, 'w') as f:
                json.dump(self.config, f)
        except Exception as e:
            logging.error(f"Error saving config: {e}")

    def setup_ui(self):
        """Initialize the user interface"""
        ctk.set_appearance_mode("light")  # Modes: system (default), light, dark
        ctk.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green
        
        self.title("GV Mahant Brothers - Invoice")
        # self.geometry(self.config["window_size"])  # Disabled to prevent window from shrinking after maximizing
        self.minsize(1200, 800)
        self.configure(padx=30, pady=30)
        
        # Maximize the window after all widgets are initialized
        self.after(100, lambda: self.state('zoomed'))

        self.current_mode = ctk.StringVar(value="Patti")
        self.rows = []
        self.row_counter = 0
        self.autosave_var = ctk.BooleanVar(value=self.config["autosave"])

        # Create tooltip label with improved styling
        self.tooltip = ttk.Label(
            self,
            background="#2c2c2c",
            foreground="white",
            relief="flat",
            borderwidth=0,
            padding=8
        )
        self.tooltip_timer = None

        self.kata_amount_entry = None

        self.numeric_vcmd = (self.register(self.only_numeric_input), '%P')
        self.check_autosave_on_start()
        self.build_ui()
        self.schedule_auto_save()

    def build_ui(self):
        # Main container frame with rounded corners and padding
        main_frame = ctk.CTkFrame(
            self,
            fg_color=BACKGROUND_COLOR,
            corner_radius=15,
            border_width=1,
            border_color=BORDER_COLOR
        )
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Header section with improved spacing
        header_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        header_frame.pack(pady=(20, 10))
        
        ctk.CTkLabel(
            header_frame,
            text="ಶ್ರೀ",
            font=HEADER_FONT,
            text_color="#213448"
        ).pack()
        
        ctk.CTkLabel(
            header_frame,
            text="G.V. Mahant Brothers",
            font=HEADER_FONT,
            text_color="#213448"
        ).pack()
        
        self.date_label = ctk.CTkLabel(
            header_frame,
            text=datetime.now().strftime("%A, %d %B %Y %I:%M:%S %p"),
            font=SUBHEADER_FONT,
            text_color="#213448"
        )
        self.date_label.pack()

        # Start updating the date/time label every second
        self.update_datetime()

        # Mode navigation with improved styling
        nav_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        nav_frame.pack(pady=20)
        
        self.mode_buttons = {}
        for i, mode in enumerate(["Patti", "Kata", "Barthe"]):
            btn = ctk.CTkButton(
                nav_frame,
                text=mode,
                command=lambda m=mode: self.set_mode(m),
                font=LABEL_FONT,
                fg_color="#547792" if self.current_mode.get() == mode else "#e0e0e0",
                text_color="white" if self.current_mode.get() == mode else "#212121",
                border_color="#547792",
                hover_color="#63889e",  # Slightly lighter version for hover
                corner_radius=20,
                width=120,
                height=38,
            )
            btn.grid(row=0, column=i, padx=10, pady=10)
            self.mode_buttons[mode] = btn

        # Customer name section with improved styling
        customer_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        customer_frame.pack(pady=20)
        
        ctk.CTkLabel(
            customer_frame,
            text="Customer Name:",
            font=LABEL_FONT,
            text_color=TEXT_COLOR
        ).pack(side="left", padx=(0, 10))
        
        self.customer_entry = ctk.CTkEntry(
            customer_frame,
            width=400,
            font=ENTRY_FONT,
            height=38,
            corner_radius=8,
            border_color=BORDER_COLOR,
            fg_color="#ffffff"
        )
        self.customer_entry.pack(side="left")

        # Create a container for the table with scrolling
        table_container = ctk.CTkFrame(main_frame, fg_color="transparent")
        table_container.pack(fill="both", expand=True, padx=20, pady=(20, 10))

        # Create a scrollable frame with automatic scrollbar
        self.scrollable_frame = ctk.CTkScrollableFrame(
            table_container,
            fg_color=FRAME_COLOR,
            corner_radius=10,
            border_width=1,
            border_color=BORDER_COLOR,
            scrollbar_button_color="#4a4a4a",  # Dark grey
            scrollbar_button_hover_color="#666666"  # Slightly lighter grey for hover
        )
        self.scrollable_frame.pack(fill="both", expand=True)

        # Set the table frame to the scrollable frame
        self.table_frame = self.scrollable_frame

        # Bottom frame with improved styling
        self.bottom_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        self.bottom_frame.pack(fill="x", pady=(10, 20), padx=20)

        # Action buttons with improved styling
        left_buttons_frame = ctk.CTkFrame(self.bottom_frame, fg_color="transparent")
        left_buttons_frame.pack(side="left", padx=(0, 20))

        button_style = {
            "font": BUTTON_FONT,
            "width": 130,
            "height": 38,
            "corner_radius": 8,
            "border_width": 1,
            "border_color": "#004d80",
            "fg_color": "#004d80",
            "hover_color": "#005c99",  # Slightly lighter version for hover
            "text_color": "white"
        }

        ctk.CTkButton(
            left_buttons_frame,
            text="Add Row",
            command=self.add_row,
            **button_style
        ).pack(side="left", padx=5)

        ctk.CTkButton(
            left_buttons_frame,
            text="Clear",
            command=self.clear_rows,
            **button_style
        ).pack(side="left", padx=5)

        ctk.CTkButton(
            left_buttons_frame,
            text="Save",
            command=self.save_to_excel_async,
            **button_style
        ).pack(side="left", padx=5)

        ctk.CTkButton(
            left_buttons_frame,
            text="Open Folder",
            command=self.open_save_folder,
            **button_style
        ).pack(side="left", padx=5)

        ctk.CTkButton(
            left_buttons_frame,
            text="Print",
            command=self.show_print_preview,
            **button_style
        ).pack(side="left", padx=5)

        # Total section with improved styling
        right_total_frame = ctk.CTkFrame(self.bottom_frame, fg_color="transparent")
        right_total_frame.pack(side="right")

        self.kata_field_frame = ctk.CTkFrame(right_total_frame, fg_color="transparent")
        self.kata_field_frame.pack(side="left", padx=(0, 15))

        self.total_label = ctk.CTkLabel(
            right_total_frame,
            text="Total Amount: ₹0.00",
            font=("Segoe UI", 24, "bold"),
            text_color="#213448"
        )
        self.total_label.pack(side="left")

        # Create initial table content
        self.create_table_headers()
        self.add_row()
        self.switch_mode()

        # Bind the canvas to update its width when the window is resized
        self.bind("<Configure>", self.on_window_resize)

    def _on_mousewheel(self, event):
        """Handle mouse wheel scrolling."""
        self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def on_window_resize(self, event):
        """Update the canvas width when the window is resized."""
        if hasattr(self, 'canvas') and hasattr(self, 'table_frame'):
            # Update the canvas window width
            self.canvas.itemconfig(self.canvas.find_withtag("all")[0], width=event.width - 100)
            # Update the table frame width
            self.table_frame.configure(width=event.width - 100)

    def create_table_headers(self):
        # Remove existing widgets
        for widget in self.table_frame.winfo_children():
            widget.destroy()

        headers = []
        mode = self.current_mode.get()
        if mode == "Patti":
            headers = ["Item", "Packet", "Quantity", "+", "Rate", "Hamali", "Amount"]
        elif mode == "Kata":
            headers = ["Item", "Net Wt", "Less%", "Final Wt", "Rate", "Hamali Rate", "Amount"]
        elif mode == "Barthe":
            headers = ["Item", "Packet", "Weight", "+", "Total Qty", "Rate", "Hamali", "Amount"]

        self._current_headers = headers

        # Create headers with improved styling
        for i, h in enumerate(headers):
            header_label = ctk.CTkLabel(
                self.table_frame,
                text=h,
                font=TABLE_HEADER_FONT,
                text_color="white",
                fg_color="#547792",  # Same as mode selection color
                corner_radius=8,
                height=38,
                anchor="center"
            )
            header_label.grid(row=0, column=i, sticky="nsew", padx=3, pady=3)
            self.table_frame.grid_columnconfigure(i, weight=1)

        # Add empty space for delete column (no header)
        self.table_frame.grid_columnconfigure(len(headers), weight=0)

    def switch_mode(self, previous_mode=None):
        """Switch between Patti, Kata, and Barthe modes with complete data isolation."""
        # 1. First capture and save the current mode data
        current_mode = previous_mode if previous_mode else self.current_mode.get()
        logging.info(f"Switching FROM {current_mode} mode")
        
        # Collect all data from current UI rows
        current_data = []
        for row_data in self.rows:
            # Skip empty rows
            if not row_data["widgets"] or not hasattr(row_data["widgets"][0], "get"):
                continue
                
            # Get the item name
            item_name = row_data["widgets"][0].get()
            if not item_name.strip():
                continue  # Skip rows with empty item names
                
            # Collect all values from this row
            row_values = []
            for widget in row_data["widgets"]:
                if isinstance(widget, (ctk.CTkEntry, ttk.Combobox)):
                    row_values.append(widget.get())
                elif isinstance(widget, ctk.CTkLabel) and "₹" in str(widget.cget("text")):
                    # Extract the amount value (remove ₹ and Error)
                    text = widget.cget("text").replace('₹', '').replace('Error', '0').strip()
                    row_values.append(text)
                else:
                    # Skip delete buttons and other non-data widgets
                    continue
            
            # Only add rows with valid data
            if len(row_values) >= 2:  # At least item name + one other field
                current_data.append(row_values)
        
        # Save current mode data if there's any content
        if current_data:
            self.mode_data[current_mode] = current_data
            self.mode_initialized[current_mode] = True
            logging.info(f"Saved {len(current_data)} rows for {current_mode} mode")
        elif self.mode_initialized[current_mode]:
            # Preserve empty data only if this mode was previously initialized with data
            self.mode_data[current_mode] = []
            logging.info(f"Saved empty data for {current_mode} mode (was previously initialized)")
        
        # 2. Clear current UI completely
        
        # Clear Kata amount field if it exists
        for widget in self.kata_field_frame.winfo_children():
            widget.destroy()
        self.kata_amount_entry = None
        
        # Clear all table rows
        for row_data in self.rows:
            for widget in row_data["widgets"]:
                widget.destroy()

        # Recreate table headers for the new mode
        self.create_table_headers()
        
        # Reset the rows list
        self.rows = []
        
        # 3. Load data for the new mode
        new_mode = self.current_mode.get()
        logging.info(f"Switching TO {new_mode} mode")
        
        # Get the saved data for this mode
        mode_specific_data = self.mode_data.get(new_mode, [])
        
        # Only add the data if this mode was previously initialized or has data
        if mode_specific_data and self.mode_initialized.get(new_mode, False):
            for row_values in mode_specific_data:
                # Skip invalid rows
                if not row_values or not row_values[0].strip():
                    continue
                
                # Create a new row
                self.add_row()
                
                # Get the widgets for the new row
                if not self.rows:
                    continue
                last_row = self.rows[-1]['widgets']
                
                # Set the item name
                if isinstance(last_row[0], ttk.Combobox):
                    last_row[0].set(row_values[0])
                
                # Set values for other fields
                field_count = min(len(row_values) - 1, len(last_row) - 2)
                for i in range(1, field_count + 1):
                    if i < len(row_values) and i < len(last_row) and hasattr(last_row[i], 'insert'):
                        last_row[i].delete(0, 'end')
                        last_row[i].insert(0, row_values[i])
        else:
            # If no data or not initialized, just add a blank row
            logging.info(f"No data found for {new_mode} mode - adding blank row")
            self.add_row()
        
        # 4. Add mode-specific UI elements
        
        # Add Kata field if needed
        if new_mode == "Kata":
            kata_label = ctk.CTkLabel(self.kata_field_frame, text="Kata:", font=LABEL_FONT)
            kata_label.pack(side="left", padx=(0, 5))
            
            self.kata_amount_entry = ctk.CTkEntry(
                self.kata_field_frame, 
                font=ENTRY_FONT,
                height=38,
                width=120,
                validate='key',
                validatecommand=self.numeric_vcmd
            )
            self.kata_amount_entry.pack(side="left")
            # Add default value '0'
            self.kata_amount_entry.insert(0, "0") 
            # Bind update on key release
            self.kata_amount_entry.bind("<KeyRelease>", self.debounce_update_amounts)
            self.kata_amount_entry.bind("<FocusIn>", self.select_all_on_focus)

        # 5. Recalculate totals
        self.update_amounts()

    def add_row(self):
        mode = self.current_mode.get()
        if mode == "Patti":
            num_entry_fields = 6
        elif mode == "Kata":
            num_entry_fields = 6
        elif mode == "Barthe":
            num_entry_fields = 7
        else:
            num_entry_fields = 5

        entries = []
        row_idx = len(self.rows) + 1

        # Item dropdown with improved styling
        item_dropdown = ttk.Combobox(
            self.table_frame,
            values=ITEM_LIST,
            font=("Segoe UI", 15),  # Increased from default 13 to 15 (15% increase)
            state="readonly"
        )
        item_dropdown.grid(row=row_idx, column=0, padx=3, pady=3, sticky="nsew")
        item_dropdown.bind("<<ComboboxSelected>>", lambda e: self.handle_item_selection(e, item_dropdown))
        self.table_frame.grid_columnconfigure(0, weight=1)
        entries.append(item_dropdown)

        # Entry fields with improved styling and numeric validation
        for i in range(1, num_entry_fields):
            entry = ctk.CTkEntry(
                self.table_frame,
                font=("Segoe UI", 15),  # Increased from default 13 to 15 (15% increase)
                justify="center",
                height=38,
                corner_radius=8,
                border_color=BORDER_COLOR,
                fg_color="#ffffff",
                validate='key',
                validatecommand=self.numeric_vcmd
            )
            entry.grid(row=row_idx, column=i, padx=3, pady=3, sticky="nsew")
            entry.bind("<KeyRelease>", self.debounce_update_amounts)
            entry.bind("<FocusIn>", self.select_all_on_focus)
            self.table_frame.grid_columnconfigure(i, weight=1)
            entries.append(entry)

        # Amount label with improved styling
        amount_label = ctk.CTkLabel(
            self.table_frame,
            text="₹0.00",
            font=("Segoe UI", 15),  # Increased from default 13 to 15 (15% increase)
            anchor="e",
            height=38,
            corner_radius=8,
            fg_color="#ffffff",
            text_color=TEXT_COLOR
        )
        amount_label.grid(row=row_idx, column=num_entry_fields, padx=3, pady=3, sticky="nsew")
        self.table_frame.grid_columnconfigure(num_entry_fields, weight=1)
        entries.append(amount_label)

        # Add delete button
        delete_btn = ctk.CTkButton(
            self.table_frame,
            text="X",
            width=40,
            height=38,
            fg_color=ERROR_COLOR,
            hover_color="#d32f2f",
            corner_radius=8,
            command=lambda: self.delete_row(row_idx)
        )
        delete_btn.grid(row=row_idx, column=num_entry_fields + 1, padx=3, pady=3)
        entries.append(delete_btn)

        self.rows.append({"row_index": row_idx, "widgets": entries})

    def handle_item_selection(self, event, dropdown):
        """Handle item selection from dropdown, including the 'Add New Item' option."""
        selected_item = dropdown.get()
        if selected_item == "Add New Item...":
            # Create a dialog for adding new item
            dialog = ctk.CTkInputDialog(
                text="Enter new item name:",
                title="Add New Item"
            )
            new_item = dialog.get_input()
            
            if new_item:
                new_item = new_item.strip().upper()
                if new_item and new_item not in ITEM_LIST[:-1]:  # Exclude "Add New Item..." from check
                    # Add the new item before "Add New Item..."
                    ITEM_LIST.insert(-1, new_item)
                    # Update all dropdowns
                    for row_data in self.rows:
                        for widget in row_data["widgets"]:
                            if isinstance(widget, ttk.Combobox):
                                widget["values"] = ITEM_LIST
                    messagebox.showinfo("Success", f"Item '{new_item}' added successfully!")
                elif new_item in ITEM_LIST[:-1]:
                    messagebox.showwarning("Warning", "This item already exists!")
                else:
                    messagebox.showwarning("Warning", "Please enter a valid item name!")
            # Reset the dropdown to empty
            dropdown.set("")
        else:
            # Normal item selection, update amounts
            self.update_amounts(event)

    def delete_row(self, row_idx):
        """Delete a specific row from the table."""
        try:
            # Don't allow deletion if only one row remains
            if len(self.rows) <= 1:
                messagebox.showwarning("Warning", "Cannot delete the last row.")
                return

            # Find and remove the row
            for i, row_data in enumerate(self.rows):
                if row_data["row_index"] == row_idx:
                    # Destroy all widgets in the row
                    for widget in row_data["widgets"]:
                        widget.destroy()
                    # Remove the row from our list
                    self.rows.pop(i)
                    break

            # Reindex remaining rows
            for i, row_data in enumerate(self.rows, 1):
                row_data["row_index"] = i
                for j, widget in enumerate(row_data["widgets"]):
                    widget.grid(row=i, column=j)

            # Update amounts after deletion
            self.update_amounts()

            # Force update the UI
            self.update_idletasks()
            
        except Exception as e:
            logging.error(f"Error deleting row: {e}")
            messagebox.showerror("Error", "Failed to delete row. Please try again.")

    def clear_rows(self):
        """Clear all rows except one."""
        try:
            # Keep only the first row
            while len(self.rows) > 1:
                row_data = self.rows[-1]
                for widget in row_data["widgets"]:
                    widget.destroy()
                self.rows.pop()

            # Reset the first row
            first_row = self.rows[0]
            for widget in first_row["widgets"]:
                if isinstance(widget, (ctk.CTkEntry, ttk.Combobox)):
                    widget.delete(0, 'end')
                elif isinstance(widget, ctk.CTkLabel):
                    widget.configure(text="₹0.00")

            # Update amounts
            self.update_amounts()
            
            # Force update the UI
            self.update_idletasks()

        except Exception as e:
            logging.error(f"Error clearing rows: {e}")
            messagebox.showerror("Error", "Failed to clear rows. Please try again.")

    def update_amounts(self, event=None):
        """Update all amount calculations in real time (no debounce)."""
        self._do_update_amounts()

    def _do_update_amounts(self):
        """Actually perform the amount updates."""
        try:
            logging.debug("Updating amounts for all rows")
            total = 0.0
            mode = self.current_mode.get()

            # Calculate sum of row amounts
            for row_data in self.rows:
                widgets = row_data["widgets"]
                amount = 0.0 # Default amount
                try:
                    if mode == "Patti":
                        # Item [0], Pkt [1], Qty [2], + [3], Rate [4], Hamali [5], AmountLabel [6]
                        if len(widgets) > 4:
                            qty = validate_float(widgets[2].get())
                            plus_value = validate_float(widgets[3].get())  # Get value from + column
                            qty = qty + plus_value  # Add the plus value to quantity
                            rate = validate_float(widgets[4].get())
                            pkt = validate_float(widgets[1].get())
                            hamali_rate = validate_float(widgets[5].get())
                            # Calculate hamali based on number of packets
                            hamali_amount = pkt * hamali_rate
                            amount = (qty * rate) - hamali_amount  # Subtract hamali amount
                    elif mode == "Kata":
                        # Item [0], Net [1], Less% [2], Final Wt [3], Rate [4], HamaliRate [5], AmountLabel [6]
                         if len(widgets) > 5:
                            net = validate_float(widgets[1].get())
                            less = validate_float(widgets[2].get())
                            final_wt = net * (1 - less / 100.0) if less < 100 else 0.0
                            # Display the calculated Final Wt
                            widgets[3].delete(0, 'end')
                            widgets[3].insert(0, f"{final_wt:.2f}")
                            rate = validate_float(widgets[4].get())
                            hamali_rate = validate_float(widgets[5].get())
                            # Calculate packets based on net weight (e.g., if 60kg/packet)
                            packets = int(net / 60) if net > 0 else 0 
                            hamali_amount = packets * hamali_rate
                            amount = (final_wt * rate) - hamali_amount  # Subtract hamali amount
                    elif mode == "Barthe":
                        # Item [0], Pkt [1], Wt   |+    |TQty |Rate |Ham |Amount |
                        # |7chars|4chr|4chr |4chr |4chr |5chr |4chr|6chars |
                         if len(widgets) > 6:
                            pkt = validate_float(widgets[1].get())
                            wt_per_pkt = validate_float(widgets[2].get())
                            adj = validate_float(widgets[3].get()) # Adjustment weight
                            total_qty = (pkt * wt_per_pkt) + adj
                            # Display the calculated Total Qty
                            widgets[4].delete(0, 'end')
                            widgets[4].insert(0, f"{total_qty:.2f}")
                            rate = validate_float(widgets[5].get()) # Rate per kg
                            hamali_rate = validate_float(widgets[6].get())
                            # Calculate hamali based on number of packets
                            hamali_amount = pkt * hamali_rate
                            amount = (total_qty * rate) - hamali_amount  # Subtract hamali amount
                    
                    # Update the amount label for the current row
                    # The amount label is always the second-to-last widget (before delete button)
                    if len(widgets) > 0:
                        widgets[-2].configure(text=f"₹{amount:.2f}")
                    total += amount

                except IndexError:
                     logging.error(f"Index error calculating amount for row. Widgets: {len(widgets)}")
                except Exception as e:
                    logging.error(f"Error calculating amount: {e}")
                    if len(widgets) > 0:
                         widgets[-2].configure(text="₹Error") # Indicate error on the row

            # --- Add Kata Amount if applicable ---
            kata_amount = 0.0
            if mode == "Kata" and self.kata_amount_entry:
                try:
                    kata_amount = validate_float(self.kata_amount_entry.get())
                    # Add visual feedback for invalid input (optional)
                    if self.kata_amount_entry.get().strip() and kata_amount == 0 and self.kata_amount_entry.get() != '0':
                         self.kata_amount_entry.configure(fg_color="pink")
                    else:
                         # Reset color on valid input
                         self.kata_amount_entry.configure(fg_color=ctk.ThemeManager.theme["CTkEntry"]["fg_color"]) 
                except Exception as e:
                    logging.error(f"Error reading Kata amount: {e}")
                    # Maybe provide visual feedback on error
                    self.kata_amount_entry.configure(fg_color="pink")
            
            total -= kata_amount # Deduct validated Kata amount from total
            # --- End Add Kata Amount ---

            self.total_label.configure(text=f"Total Amount: ₹{total:.2f}")

        except Exception as e:
            error_msg = f"Error updating amounts: {str(e)}"
            logging.error(error_msg)
            self.total_label.configure(text="₹Error")

    def save_to_excel(self, show_popup=True, filename=None):
        try:
            # Define the save directory
            save_dir = INVOICE_SAVE_DIR
        
            # Ensure the save directory exists, create if not
            os.makedirs(save_dir, exist_ok=True) 
        
            # Determine the filename
            if filename is not None:
                # If filename is provided, assume it's just the name, not full path yet
                full_save_path = os.path.join(save_dir, filename) 
            else:
                # Create the filename based on the current date
                date_str = datetime.now().strftime('%Y-%m-%d')
                base_filename = f"Invoice_{date_str}.xlsx"
                full_save_path = os.path.join(save_dir, base_filename)
            
            logging.info(f"Target save path: {full_save_path}")

            # Get Invoice Data
            customer = self.customer_entry.get().strip() or "Unknown Customer"
            mode = self.current_mode.get()
            
            headers = getattr(self, '_current_headers', []) 
            if not headers:
                if mode == "Patti":
                    headers = ["Item", "Packet", "Quantity", "+", "Rate", "Hamali", "Amount"]
                elif mode == "Kata":
                    headers = ["Item", "Net Wt", "Less%", "Rate", "Hamali Rate", "Amount"]
                elif mode == "Barthe":
                    headers = ["Item", "Packet", "Weight", "+/-", "Rate", "Hamali", "Amount"]
                else: 
                    headers = ["Col1", "Col2", "Col3", "Col4", "Col5", "Col6", "Amount"]

            data_rows = []
            total_invoice_amount = 0.0
            kata_amount = 0.0
            if mode == "Kata" and self.kata_amount_entry:
                kata_amount = validate_float(self.kata_amount_entry.get())
            for row_data in self.rows:
                widgets = row_data["widgets"]
                row_values = []
                for w in widgets:
                    if isinstance(w, (ctk.CTkEntry, ttk.Combobox)):
                        row_values.append(w.get())
                    elif isinstance(w, ctk.CTkLabel):
                        text = w.cget("text").replace('₹', '').replace('Error', '0').strip()
                        row_values.append(text)
                    else:
                        row_values.append("")

                if not row_values or not row_values[0]:
                    continue
            
                try:
                    if mode == "Patti" and len(row_values) >= 6:
                        # Calculate hamali amount
                        pkt = float(row_values[1].replace('Error', '0').strip() or 0)
                        qty = float(row_values[2].replace('Error', '0').strip() or 0)
                        plus_value = float(row_values[3].replace('Error', '0').strip() or 0)
                        total_qty = qty + plus_value
                        rate = float(row_values[4].replace('Error', '0').strip() or 0)
                        hamali_rate = float(row_values[5].replace('Error', '0').strip() or 0)
                        hamali_amount = int(pkt * hamali_rate)
                        amount_val = float(row_values[6].replace('₹', '').replace('Error', '0').strip() or 0)
                        amount_str = f"{amount_val:.2f}"
                        data_rows.append(row_values)
                        total_invoice_amount += amount_val
                    elif mode == "Kata" and len(row_values) >= 7:
                        # Calculate hamali amount
                        net = float(row_values[1].replace('Error', '0').strip() or 0)
                        packets = int(net / 60) if net > 0 else 0
                        hamali_rate = float(row_values[5].replace('Error', '0').strip() or 0)
                        hamali_amount = int(packets * hamali_rate)
                        amount_val = float(row_values[6].replace('₹', '').replace('Error', '0').strip() or 0)
                        amount_str = f"{amount_val:.2f}"
                        data_rows.append(row_values)
                        total_invoice_amount += amount_val
                    elif mode == "Barthe" and len(row_values) >= 8:
                        # Calculate hamali amount
                        pkt = float(row_values[1].replace('Error', '0').strip() or 0)
                        ham_rate = float(row_values[6].replace('Error', '0').strip() or 0)
                        ham_amount = pkt * ham_rate
                        amount_val = float(row_values[7].replace('₹', '').replace('Error', '0').strip() or 0)
                        amount_str = f"{amount_val:.2f}"
                        data_rows.append(row_values)
                        total_invoice_amount += amount_val
                except Exception as fmt_e:
                    logging.error(f"Error formatting row: {fmt_e}")

            if not data_rows:
                if show_popup:
                    messagebox.showwarning("No Data", "No data entered to save.")
                return

            # Excel Writing Logic with proper workbook handling
            try:
                if os.path.exists(full_save_path):
                    wb = load_workbook(full_save_path)
                else:
                    wb = Workbook()
                    
                # Check if mode sheet exists, create or get it
                if mode in wb.sheetnames:
                    ws = wb[mode]
                else:
                    if len(wb.sheetnames) > 0:
                        # If there are sheets but none match our mode, create new one
                        ws = wb.create_sheet(title=mode)
                    else:
                        # If it's a new workbook, rename the default sheet
                        ws = wb.active
                        ws.title = mode

                # Only write headers if the first row is empty or not matching
                expected_headers = ["Timestamp", "Customer"] + headers
                first_row = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
                if first_row != expected_headers:
                    ws.insert_rows(1)
                    for col, value in enumerate(expected_headers, 1):
                        ws.cell(row=1, column=col, value=value)

                # Write data
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                for row in data_rows:
                    ws.append([timestamp, customer] + row)

                # --- Attempt to save the file --- 
                primary_save_path = full_save_path
                desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
                fallback_save_path = os.path.join(desktop_path, os.path.basename(primary_save_path))

                try:
                    wb.save(primary_save_path)
                    logging.info(f"Successfully saved invoice data to {primary_save_path} (Sheet: {mode})")
                    if show_popup:
                        messagebox.showinfo("Saved", f"Invoice data saved to:\n{primary_save_path}\n(Sheet: {mode})")

                except (PermissionError, OSError, IOError) as e_primary:
                    logging.warning(f"Failed to save to primary path {primary_save_path}: {e_primary}. Attempting fallback to Desktop.")
                    try:
                        # Ensure fallback directory exists (Desktop usually does, but good practice)
                        os.makedirs(desktop_path, exist_ok=True)
                        wb.save(fallback_save_path)
                        logging.info(f"Successfully saved invoice data to fallback path: {fallback_save_path} (Sheet: {mode})")
                        if show_popup:
                            messagebox.showinfo("Saved to Desktop", f"Could not save to D:\\invoices.\nFile saved to Desktop instead:\n{fallback_save_path}\n(Sheet: {mode})")
                    except Exception as e_fallback:
                        error_msg = f"Failed to save to both primary location and Desktop.\nPrimary Error: {e_primary}\nFallback Error: {e_fallback}"
                        logging.error(error_msg)
                        if show_popup:
                            messagebox.showerror("Save Error", error_msg)

            except Exception as e:
                error_msg = f"Error saving Excel file to:\n{full_save_path}\n\nError: {str(e)}"
                logging.error(error_msg)
                if show_popup:
                    messagebox.showerror("Save Error", error_msg)

        except Exception as e:
            error_msg = f"Unexpected error during save operation: {str(e)}"
            logging.exception(error_msg)
            if show_popup:
                messagebox.showerror("Error", error_msg)

    def format_line(self, left, right, width=42):
        space = width - len(left) - len(right)
        return f"{left}{' ' * max(space, 0)}{right}"

    def generate_print_content(self):
        lines = []
        customer = self.customer_entry.get().strip() or "N/A"
        mode = self.current_mode.get()
        max_width = 48  # updated width for your printer

        # Header
        lines.append("G.V. Mahant Brothers".center(max_width))
        lines.append(datetime.now().strftime("%d-%b-%Y %H:%M").center(max_width))
        # Add 'Customer Name: <name>' centered
        lines.append(f"Customer Name: {customer}".center(max_width))
        lines.append("-" * max_width)
        # Define and print header before looping rows
        if mode == "Patti":
            header_fmt = "{:<8} {:>4} {:>5} {:>7} {:>6}{:>14}"
            lines.append(header_fmt.format("Item", "Pkt", "Qty", "Rate", "Hm", "Amount"))
        elif mode == "Kata":
            # Narrower Item column, Net closer, Amount inside border
            header_fmt = "{:<9}{:>6} {:>7}{:>6}{:>6}{:>10}"
            lines.append(header_fmt.format("Item", "Net", "FWt", "Rt", "Hm", "Amount"))
        elif mode == "Barthe":
            header_fmt = "{:<8}{:>5}{:>6}{:>8}{:>5}{:>6}{:>10}"
            lines.append(header_fmt.format("Item", "Pkt", "Wt", "TQty", "Rt", "Hm", "Amount"))
        else:
            header_fmt = None
            lines.append("Unknown mode".center(max_width))
        lines.append("-" * max_width)

        total = 0.0
        if header_fmt:
            for row_data in self.rows:
                widgets = row_data["widgets"]
                row_values = []
                for w in widgets:
                    if isinstance(w, (ctk.CTkEntry, ttk.Combobox)):
                        row_values.append(w.get())
                    elif isinstance(w, ctk.CTkLabel):
                        text = w.cget("text").replace('₹', '').replace('Error', '0').strip()
                        row_values.append(text)
                    else:
                        row_values.append("")

                if not row_values or not row_values[0]:
                    continue
                values = [w.get().strip() if hasattr(w, 'get') else "" for w in widgets]
                amount_label = widgets[-2]
                amt = amount_label.cget("text").replace("₹", "").strip()
                try:
                    amt_val = float(amt)
                    total += amt_val
                except:
                    amt_val = 0.0
                if mode == "Patti" and len(values) >= 6:
                    try:
                        pkt = float(values[1] or 0)
                        hamali_rate = float(values[5] or 0)
                        ham_amount = pkt * hamali_rate
                    except:
                        ham_amount = 0
                    lines.append(header_fmt.format(
                        values[0][:8],   # Item (truncate to 8)
                        values[1],       # Pkt
                        values[2],       # Qty
                        values[4],       # Rate
                        f"{ham_amount:.0f}", # Hm (calculated)
                        f"{amt_val:.2f}" # Amount
                    ))
                elif mode == "Kata" and len(values) >= 7:
                    try:
                        net = float(values[1] or 0)
                        final_wt = float(values[3] or 0)
                        rate = float(values[4] or 0)
                        hamali_rate = float(values[5] or 0)
                        # Calculate packets as int(net / 60)
                        packets = int(net / 60) if net > 0 else 0
                        ham_amount = packets * hamali_rate
                    except:
                        net = final_wt = rate = hamali_rate = packets = ham_amount = 0
                    lines.append(header_fmt.format(
                        values[0][:9],   # Item (truncate to 9)
                        values[1],        # Net
                        values[3],        # FWt
                        values[4],        # Rt
                        f"{ham_amount:.0f}", # Hm (calculated value)
                        f"{amt_val:.2f}"  # Amount (inside the lines)
                    ))
                elif mode == "Barthe" and len(values) >= 8:
                    try:
                        pkt = float(values[1] or 0)
                        wt = float(values[2] or 0)
                        tqty = float(values[4] or 0)
                        rate = float(values[5] or 0)
                        ham_rate = float(values[6] or 0)
                        ham_amount = pkt * ham_rate
                    except:
                        pkt = wt = tqty = rate = ham_rate = ham_amount = 0
                    lines.append(header_fmt.format(
                        values[0][:8],   # Item (truncate to 8)
                        values[1],       # Pkt
                        values[2],       # Wt
                        f"{tqty:.2f}",   # TQty
                        values[5],       # Rt
                        f"{ham_amount:.0f}", # Hm
                        f"{amt_val:.2f}" # Amt
                    ))

        if mode == "Kata" and hasattr(self, 'kata_amount_entry') and self.kata_amount_entry:
            kata_val = validate_float(self.kata_amount_entry.get())
            total -= kata_val
            kata_line = f"    Kata Amount:{kata_val:>30.2f}"
            lines.append(kata_line)

        lines.append("-" * max_width)
        # Indent and center Total Amount
        total_line = f"Total Amount: {total:.2f}"
        lines.append(total_line.center(max_width))
        lines.append("-" * max_width)
        lines.extend(["\n"] * 3)
        lines.append(chr(27) + chr(105))  # Cut command

        return lines

    def save_for_print(self):
        """Prints the generated content to the default printer."""
        try:
            printer_name = win32print.GetDefaultPrinter()
            logging.info(f"Attempting to print to default printer: {printer_name}")
            
            lines = self.generate_print_content()
            print_content = "\r\n".join(lines)
            
            # First try UTF-8 encoding for Kannada text
            try:
                print_bytes = print_content.encode('utf-8')
            except UnicodeEncodeError:
                # If UTF-8 fails, try UTF-16
                try:
                    print_bytes = print_content.encode('utf-16')
                except UnicodeEncodeError:
                    # If both fail, fall back to cp437 with replacement
                    print_bytes = print_content.encode('cp437', errors='replace')

            # Use win32print for direct RAW printing
            hPrinter = win32print.OpenPrinter(printer_name)
            try:
                # Job name "Invoice", Datatype "RAW"
                hJob = win32print.StartDocPrinter(hPrinter, 1, ("Invoice", None, "RAW"))
                try:
                    win32print.StartPagePrinter(hPrinter)
                    win32print.WritePrinter(hPrinter, print_bytes)
                    win32print.EndPagePrinter(hPrinter)
                finally:
                    win32print.EndDocPrinter(hPrinter)
            finally:
                win32print.ClosePrinter(hPrinter)

            logging.info("Invoice successfully sent to printer.")
            messagebox.showinfo("Success", "Invoice sent to printer!")

        except Exception as e:
            error_msg = f"Error printing invoice: {str(e)}"
            logging.error(error_msg)
            logging.error("Printer encoding error - trying to print Kannada text")
            messagebox.showerror("Print Error", f"Could not print to {printer_name}.\nCheck if your printer supports Kannada text.\n\nError: {e}")


    def show_print_preview(self):
        """Shows a Toplevel window with a preview of the print output."""
        try:
            # Auto-save before showing preview
            self.save_to_excel(show_popup=False)
            
            preview = ctk.CTkToplevel(self)
            preview.title("Print Preview")
            preview.geometry("450x600") # Slightly wider for better view
            preview.transient(self) # Keep preview on top of main window
            preview.grab_set()  # Make the window modal

            # Center the preview window relative to the main app
            app_x = self.winfo_x()
            app_y = self.winfo_y()
            app_w = self.winfo_width()
            app_h = self.winfo_height()
            pre_w = 450
            pre_h = 600
            x = app_x + (app_w - pre_w) // 2
            y = app_y + (app_h - pre_h) // 2
            preview.geometry(f"{pre_w}x{pre_h}+{x}+{y}")

            # --- Preview Content Area ---
            # Use a CTkTextbox within a ScrollableFrame for better handling
            scroll_frame = ctk.CTkScrollableFrame(preview, label_text="Preview")
            scroll_frame.pack(fill="both", expand=True, padx=10, pady=(10, 0))

            preview_text = ctk.CTkTextbox(
                scroll_frame, 
                font=("Courier New", 10), # Monospace font for alignment
                wrap="none" # Prevent wrapping to see true line breaks
            )
            preview_text.pack(fill="both", expand=True)
            
            # Generate print content and display it
            lines = self.generate_print_content()
            # Join lines, but remove the final cut command for preview
            preview_content = "\n".join(lines[:-1]) if lines else "" 
            
            preview_text.insert("1.0", preview_content)
            preview_text.configure(state="disabled")  # Make read-only

            # --- Buttons Frame ---
            button_frame = ctk.CTkFrame(preview, fg_color="transparent")
            button_frame.pack(fill="x", padx=10, pady=10)
            
            # Center buttons using grid
            button_frame.grid_columnconfigure(0, weight=1)
            button_frame.grid_columnconfigure(1, weight=1)

            ctk.CTkButton(
                button_frame,
                text="Print",
                # Lambda calls destroy first, then the print function
                command=lambda: [preview.destroy(), self.save_for_print()], 
                width=120
            ).grid(row=0, column=0, padx=5, pady=5, sticky="ew")

            ctk.CTkButton(
                button_frame,
                text="Close",
                command=preview.destroy,
                width=120
            ).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
            
            preview.after(100, preview.lift) # Ensure it's raised above main window

        except Exception as e:
            error_msg = f"Error generating print preview: {str(e)}"
            logging.error(error_msg)
            messagebox.showerror("Preview Error", error_msg)
            if 'preview' in locals() and preview.winfo_exists():
                 preview.destroy() # Close broken preview window

    def update_datetime(self):
        """Update the date/time label with current time."""
        # Update every 5 seconds instead of every second
        current_time = datetime.now().strftime("%I:%M:%S %p")
        current_date = datetime.now().strftime("%A, %d %B %Y")
        if not hasattr(self, '_last_date') or self._last_date != current_date:
            self._last_date = current_date
            self.date_label.configure(text=f"{current_date} {current_time}")
        else:
            # Only update the time portion
            self.date_label.configure(text=f"{self._last_date} {current_time}")
        self.after(5000, self.update_datetime)  # Update every 5 seconds

    def set_mode(self, mode):
        """Set the current mode and update button colors."""
        previous_mode = self.current_mode.get()  # Save old mode first
        self.current_mode.set(mode)  # Then update to the new mode
        
        # Update button colors for all modes
        for btn_mode, btn in self.mode_buttons.items():
            if btn_mode == mode:
                btn.configure(
                    fg_color="#547792",
                    text_color="white",
                    border_color="#547792",
                    hover_color="#63889e"
                )
            else:
                btn.configure(
                    fg_color="#e0e0e0",
                    text_color="#212121",
                    border_color="#d0d0d0",
                    hover_color="#d0d0d0"
                )
        
        # Pass previous_mode to switch_mode
        self.switch_mode(previous_mode)

    def only_numeric_input(self, P):
        # Allow empty string, integer, or float
        if P == "" or (P.replace('.', '', 1).isdigit() and P.count('.') <= 1):
            return True
        return False

    def debounce_update_amounts(self, event=None):
        if hasattr(self, '_debounce_after_id') and self._debounce_after_id:
            self.after_cancel(self._debounce_after_id)
        self._debounce_after_id = self.after(300, self._do_update_amounts)  # 300ms debounce

    def select_all_on_focus(self, event):
        event.widget.select_range(0, 'end')
        event.widget.icursor('end')

    def open_save_folder(self):
        """Opens the folder where invoices are saved."""
        save_dir = INVOICE_SAVE_DIR
        try:
            if os.path.exists(save_dir):
                os.startfile(save_dir) # Opens the folder in Windows Explorer
                logging.info(f"Opened save folder: {save_dir}")
            else:
                logging.warning(f"Save folder not found: {save_dir}")
                # Try to create it? Or just inform user?
                os.makedirs(save_dir, exist_ok=True) # Attempt to create if missing
                if os.path.exists(save_dir):
                     messagebox.showinfo("Folder Created", f"The save folder ({save_dir}) was created.")
                     os.startfile(save_dir)
                else:
                    messagebox.showwarning("Folder Not Found", f"The save folder ({save_dir}) could not be found or created.")
        except Exception as e:
            logging.error(f"Error opening folder {save_dir}: {e}")
            messagebox.showerror("Error", f"Could not open the folder.\nError: {e}")

    def save_to_excel_async(self):
        threading.Thread(target=self.save_to_excel, daemon=True).start()

    def auto_save(self):
        # Save to a special autosave file
        self.save_to_excel(filename='autosave_invoice.xlsx', show_popup=False)

    def check_autosave_on_start(self):
        autosave_path = os.path.join(INVOICE_SAVE_DIR, 'autosave_invoice.xlsx')
        if os.path.exists(autosave_path):
            if messagebox.askyesno("Recover?", f"Recover unsaved invoice from last session?\n({autosave_path})"):
                self.load_invoice(autosave_path)

    def load_invoice(self, filename):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename)
            ws = wb.active
            # Clear current rows
            self.clear_rows()
            # Read the first data row (skip header)
            first_data_row = None
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if i == 0:
                    # Set customer name from first row
                    self.customer_entry.delete(0, 'end')
                    self.customer_entry.insert(0, row[1])
                # Add a row and fill in values
                self.add_row()
                last_row = self.rows[-1]['widgets']
                # Fill in item and entry fields (skip timestamp, customer, and amount)
                for j, value in enumerate(row[2:-1]):
                    if hasattr(last_row[j], 'delete'):
                        last_row[j].delete(0, 'end')
                        last_row[j].insert(0, str(value) if value is not None else "")
            self.update_amounts()
            import os
            os.remove(filename)  # Remove autosave after recovery
            messagebox.showinfo('Recovered', 'Invoice data recovered from autosave.')
        except Exception as e:
            messagebox.showerror('Error', f'Could not load invoice: {e}')

    def schedule_auto_save(self):
        self.auto_save()
        self.after(900000, self.schedule_auto_save)  # 15 minutes (900,000 ms)

if __name__ == "__main__":
    app = InvoiceApp()
    app.mainloop()