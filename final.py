import customtkinter as ctk
from tkinter import messagebox, ttk, Toplevel, Text, Scrollbar
from datetime import datetime
import os
import logging
from openpyxl import Workbook, load_workbook
import json
import win32print
import win32api
from win32printing import Printer
import codecs
import threading

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Constants
CONFIG_FILE = "app_config.json"
AUTOSAVE_INTERVAL = 300000  # 5 minutes in milliseconds

# Item list for dropdown
ITEM_LIST = [
    "MAIZE", "SOYABEAN", "LOBHA", "HULLI", "KADLI", "BLACK MOONG", 
    "CHAMAKI MOONG", "RAGI", "WHEAT", "RICE", "BILAJOLA", "BIJAPUR", 
    "CHS-5", "FEEDS", "KUSUBI", "SASAVI", "SAVI", "CASTER SEEDS", 
    "TOOR RED", "TOOR WHITE", "HUNASIBIKA", "SF", "AWARI",
    "Add New Item..."  # Add this as the last option
]

def validate_float(value):
    """Validate if a string can be converted to float."""
    try:
        return float(value) if value.strip() else 0
    except ValueError:
        return 0

# Define font configurations
HEADER_FONT = ("Segoe UI", 28, "bold")
SUBHEADER_FONT = ("Segoe UI", 16)
LABEL_FONT = ("Segoe UI", 13)
ENTRY_FONT = ("Segoe UI", 13)
TABLE_HEADER_FONT = ("Segoe UI", 13, "bold")
TABLE_FONT = ("Segoe UI", 13)
BUTTON_FONT = ("Segoe UI", 13)

# Define color scheme for light theme
PRIMARY_COLOR = "#1976d2"      # Blue
SECONDARY_COLOR = "#2196f3"    # Lighter blue
ACCENT_COLOR = "#64b5f6"       # Even lighter blue
BACKGROUND_COLOR = "#ffffff"    # White
FRAME_COLOR = "#f5f5f5"        # Light gray
BORDER_COLOR = "#e0e0e0"       # Border gray
TEXT_COLOR = "#212121"         # Dark gray for text
ERROR_COLOR = "#f44336"        # Red

class InvoiceApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.load_config()
        self.setup_ui()
        # Removed the call to self.schedule_autosave() since it's not defined
        # Uncomment the next line if you plan to use autosave later
        # self.schedule_autosave()
        self._update_timer = None  # Add this line for debouncing

    def load_config(self):
        """Load application configuration from file"""
        self.config = {
            "theme": "Green",
            "window_size": "1200x800",
            "autosave": True
        }
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, 'r') as f:
                    self.config.update(json.load(f))
        except Exception as e:
            logging.error(f"Error loading config: {e}")

    def save_config(self):
        """Save application configuration to file"""
        try:
            with open(CONFIG_FILE, 'w') as f:
                json.dump(self.config, f)
        except Exception as e:
            logging.error(f"Error saving config: {e}")

    def setup_ui(self):
        """Initialize the user interface"""
        ctk.set_appearance_mode("light")  # Modes: system (default), light, dark
        ctk.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green
        
        self.title("GV Mahant Brothers - Invoice")
        # self.geometry(self.config["window_size"])  # Disabled to prevent window from shrinking after maximizing
        self.minsize(1200, 800)
        self.configure(padx=30, pady=30)
        
        # Maximize the window after all widgets are initialized
        self.after(100, lambda: self.state('zoomed'))

        self.current_mode = ctk.StringVar(value="Patti")
        self.rows = []
        self.row_counter = 0
        self.autosave_var = ctk.BooleanVar(value=self.config["autosave"])

        # Create tooltip label with improved styling
        self.tooltip = ttk.Label(
            self,
            background="#2c2c2c",
            foreground="white",
            relief="flat",
            borderwidth=0,
            padding=8
        )
        self.tooltip_timer = None

        self.kata_amount_entry = None

        self.numeric_vcmd = (self.register(self.only_numeric_input), '%P')
        self.check_autosave_on_start()
        self.build_ui()
        self.schedule_auto_save()

    def build_ui(self):
        # Main container frame with rounded corners and padding
        main_frame = ctk.CTkFrame(
            self,
            fg_color=BACKGROUND_COLOR,
            corner_radius=15,
            border_width=1,
            border_color=BORDER_COLOR
        )
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Header section with improved spacing
        header_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        header_frame.pack(pady=(20, 10))
        
        ctk.CTkLabel(
            header_frame,
            text="ಶ್ರೀ",
            font=HEADER_FONT,
            text_color="#213448"
        ).pack()
        
        ctk.CTkLabel(
            header_frame,
            text="G.V. Mahant Brothers",
            font=HEADER_FONT,
            text_color="#213448"
        ).pack()
        
        self.date_label = ctk.CTkLabel(
            header_frame,
            text=datetime.now().strftime("%A, %d %B %Y %I:%M:%S %p"),
            font=SUBHEADER_FONT,
            text_color="#213448"
        )
        self.date_label.pack()

        # Start updating the date/time label every second
        self.update_datetime()

        # Mode navigation with improved styling
        nav_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        nav_frame.pack(pady=20)
        
        self.mode_buttons = {}
        for i, mode in enumerate(["Patti", "Kata", "Barthe"]):
            btn = ctk.CTkButton(
                nav_frame,
                text=mode,
                command=lambda m=mode: self.set_mode(m),
                font=LABEL_FONT,
                fg_color="#547792" if self.current_mode.get() == mode else "#e0e0e0",
                text_color="white" if self.current_mode.get() == mode else "#212121",
                border_color="#547792",
                hover_color="#63889e",  # Slightly lighter version for hover
                corner_radius=20,
                width=120,
                height=38,
            )
            btn.grid(row=0, column=i, padx=10, pady=10)
            self.mode_buttons[mode] = btn

        # Customer name section with improved styling
        customer_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        customer_frame.pack(pady=20)
        
        ctk.CTkLabel(
            customer_frame,
            text="Customer Name:",
            font=LABEL_FONT,
            text_color=TEXT_COLOR
        ).pack(side="left", padx=(0, 10))
        
        self.customer_entry = ctk.CTkEntry(
            customer_frame,
            width=400,
            font=ENTRY_FONT,
            height=38,
            corner_radius=8,
            border_color=BORDER_COLOR,
            fg_color="#ffffff"
        )
        self.customer_entry.pack(side="left")

        # Create a container for the table with scrolling
        table_container = ctk.CTkFrame(main_frame, fg_color="transparent")
        table_container.pack(fill="both", expand=True, padx=20, pady=(20, 10))

        # Create a scrollable frame with automatic scrollbar
        self.scrollable_frame = ctk.CTkScrollableFrame(
            table_container,
            fg_color=FRAME_COLOR,
            corner_radius=10,
            border_width=1,
            border_color=BORDER_COLOR,
            scrollbar_button_color="#4a4a4a",  # Dark grey
            scrollbar_button_hover_color="#666666"  # Slightly lighter grey for hover
        )
        self.scrollable_frame.pack(fill="both", expand=True)

        # Set the table frame to the scrollable frame
        self.table_frame = self.scrollable_frame

        # Bottom frame with improved styling
        self.bottom_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        self.bottom_frame.pack(fill="x", pady=(10, 20), padx=20)

        # Action buttons with improved styling
        left_buttons_frame = ctk.CTkFrame(self.bottom_frame, fg_color="transparent")
        left_buttons_frame.pack(side="left", padx=(0, 20))

        button_style = {
            "font": BUTTON_FONT,
            "width": 130,
            "height": 38,
            "corner_radius": 8,
            "border_width": 1,
            "border_color": "#004d80",
            "fg_color": "#004d80",
            "hover_color": "#005c99",  # Slightly lighter version for hover
            "text_color": "white"
        }

        ctk.CTkButton(
            left_buttons_frame,
            text="Add Row",
            command=self.add_row,
            **button_style
        ).pack(side="left", padx=5)

        ctk.CTkButton(
            left_buttons_frame,
            text="Clear",
            command=self.clear_rows,
            **button_style
        ).pack(side="left", padx=5)

        ctk.CTkButton(
            left_buttons_frame,
            text="Save",
            command=self.save_to_excel_async,
            **button_style
        ).pack(side="left", padx=5)

        ctk.CTkButton(
            left_buttons_frame,
            text="Print",
            command=self.show_print_preview,
            **button_style
        ).pack(side="left", padx=5)

        # Total section with improved styling
        right_total_frame = ctk.CTkFrame(self.bottom_frame, fg_color="transparent")
        right_total_frame.pack(side="right")

        self.kata_field_frame = ctk.CTkFrame(right_total_frame, fg_color="transparent")
        self.kata_field_frame.pack(side="left", padx=(0, 15))

        self.total_label = ctk.CTkLabel(
            right_total_frame,
            text="Total Amount: ₹0.00",
            font=("Segoe UI", 24, "bold"),
            text_color="#213448"
        )
        self.total_label.pack(side="left")

        # Create initial table content
        self.create_table_headers()
        self.add_row()
        self.switch_mode()

        # Bind the canvas to update its width when the window is resized
        self.bind("<Configure>", self.on_window_resize)

    def _on_mousewheel(self, event):
        """Handle mouse wheel scrolling."""
        self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def on_window_resize(self, event):
        """Update the canvas width when the window is resized."""
        if hasattr(self, 'canvas') and hasattr(self, 'table_frame'):
            # Update the canvas window width
            self.canvas.itemconfig(self.canvas.find_withtag("all")[0], width=event.width - 100)
            # Update the table frame width
            self.table_frame.configure(width=event.width - 100)

    def create_table_headers(self):
        # Remove existing widgets
        for widget in self.table_frame.winfo_children():
            widget.destroy()

        headers = []
        mode = self.current_mode.get()
        if mode == "Patti":
            headers = ["Item", "Packet", "Quantity", "Rate", "Hamali", "Amount"]
        elif mode == "Kata":
            headers = ["Item", "Net Wt", "Less%", "Rate", "Hamali Rate", "Amount"]
        elif mode == "Barthe":
            headers = ["Item", "Packet", "Weight", "+/-", "Rate", "Hamali", "Amount"]

        self._current_headers = headers

        # Create headers with improved styling
        for i, h in enumerate(headers):
            header_label = ctk.CTkLabel(
                self.table_frame,
                text=h,
                font=TABLE_HEADER_FONT,
                text_color="white",
                fg_color="#547792",  # Same as mode selection color
                corner_radius=8,
                height=38,
                anchor="center"
            )
            header_label.grid(row=0, column=i, sticky="nsew", padx=3, pady=3)
            self.table_frame.grid_columnconfigure(i, weight=1)

        # Add empty space for delete column (no header)
        self.table_frame.grid_columnconfigure(len(headers), weight=0)

    def switch_mode(self):
        # --- Clear Kata field if it exists ---
        for widget in self.kata_field_frame.winfo_children():
            widget.destroy()
        self.kata_amount_entry = None # Reset the variable
        # --- End Clear Kata field ---

        # Recreate headers which also clears the table frame's rows
        self.create_table_headers()
        # Clear the logical rows list
        self.rows.clear() 
        self.add_row() # Add a new blank row for the new mode

        # --- Add Kata field if mode is Kata ---
        if self.current_mode.get() == "Kata":
            kata_label = ctk.CTkLabel(self.kata_field_frame, text="Kata:", font=LABEL_FONT)
            kata_label.pack(side="left", padx=(0, 5))
            
            self.kata_amount_entry = ctk.CTkEntry(
                self.kata_field_frame, 
                font=ENTRY_FONT,
                height=38,
                width=120,
                validate='key',
                validatecommand=self.numeric_vcmd
            )
            self.kata_amount_entry.pack(side="left")
            # Add default value '0'
            self.kata_amount_entry.insert(0, "0") 
            # Bind update on key release
            self.kata_amount_entry.bind("<KeyRelease>", self.debounce_update_amounts)
            self.kata_amount_entry.bind("<FocusIn>", self.select_all_on_focus)
        # --- End Add Kata field ---

        self.update_amounts() # Recalculate total

    def add_row(self):
        mode = self.current_mode.get()
        if mode == "Patti":
            num_entry_fields = 5
        elif mode == "Kata":
            num_entry_fields = 5
        elif mode == "Barthe":
            num_entry_fields = 6
        else:
            num_entry_fields = 5

        entries = []
        row_idx = len(self.rows) + 1

        # Item dropdown with improved styling
        item_dropdown = ttk.Combobox(
            self.table_frame,
            values=ITEM_LIST,
            font=("Segoe UI", 15),  # Increased from default 13 to 15 (15% increase)
            state="readonly"
        )
        item_dropdown.grid(row=row_idx, column=0, padx=3, pady=3, sticky="nsew")
        item_dropdown.bind("<<ComboboxSelected>>", lambda e: self.handle_item_selection(e, item_dropdown))
        self.table_frame.grid_columnconfigure(0, weight=1)
        entries.append(item_dropdown)

        # Entry fields with improved styling and numeric validation
        for i in range(1, num_entry_fields):
            entry = ctk.CTkEntry(
                self.table_frame,
                font=("Segoe UI", 15),  # Increased from default 13 to 15 (15% increase)
                justify="center",
                height=38,
                corner_radius=8,
                border_color=BORDER_COLOR,
                fg_color="#ffffff",
                validate='key',
                validatecommand=self.numeric_vcmd
            )
            entry.grid(row=row_idx, column=i, padx=3, pady=3, sticky="nsew")
            entry.bind("<KeyRelease>", self.debounce_update_amounts)
            entry.bind("<FocusIn>", self.select_all_on_focus)
            self.table_frame.grid_columnconfigure(i, weight=1)
            entries.append(entry)

        # Amount label with improved styling
        amount_label = ctk.CTkLabel(
            self.table_frame,
            text="₹0.00",
            font=("Segoe UI", 15),  # Increased from default 13 to 15 (15% increase)
            anchor="e",
            height=38,
            corner_radius=8,
            fg_color="#ffffff",
            text_color=TEXT_COLOR
        )
        amount_label.grid(row=row_idx, column=num_entry_fields, padx=3, pady=3, sticky="nsew")
        self.table_frame.grid_columnconfigure(num_entry_fields, weight=1)
        entries.append(amount_label)

        # Add delete button
        delete_btn = ctk.CTkButton(
            self.table_frame,
            text="X",
            width=40,
            height=38,
            fg_color=ERROR_COLOR,
            hover_color="#d32f2f",
            corner_radius=8,
            command=lambda: self.delete_row(row_idx)
        )
        delete_btn.grid(row=row_idx, column=num_entry_fields + 1, padx=3, pady=3)
        entries.append(delete_btn)

        self.rows.append({"row_index": row_idx, "widgets": entries})

    def handle_item_selection(self, event, dropdown):
        """Handle item selection from dropdown, including the 'Add New Item' option."""
        selected_item = dropdown.get()
        if selected_item == "Add New Item...":
            # Create a dialog for adding new item
            dialog = ctk.CTkInputDialog(
                text="Enter new item name:",
                title="Add New Item"
            )
            new_item = dialog.get_input()
            
            if new_item:
                new_item = new_item.strip().upper()
                if new_item and new_item not in ITEM_LIST[:-1]:  # Exclude "Add New Item..." from check
                    # Add the new item before "Add New Item..."
                    ITEM_LIST.insert(-1, new_item)
                    # Update all dropdowns
                    for row_data in self.rows:
                        for widget in row_data["widgets"]:
                            if isinstance(widget, ttk.Combobox):
                                widget["values"] = ITEM_LIST
                    messagebox.showinfo("Success", f"Item '{new_item}' added successfully!")
                elif new_item in ITEM_LIST[:-1]:
                    messagebox.showwarning("Warning", "This item already exists!")
                else:
                    messagebox.showwarning("Warning", "Please enter a valid item name!")
            # Reset the dropdown to empty
            dropdown.set("")
        else:
            # Normal item selection, update amounts
            self.update_amounts(event)

    def delete_row(self, row_idx):
        """Delete a specific row from the table."""
        try:
            # Don't allow deletion if only one row remains
            if len(self.rows) <= 1:
                messagebox.showwarning("Warning", "Cannot delete the last row.")
                return

            # Find and remove the row
            for i, row_data in enumerate(self.rows):
                if row_data["row_index"] == row_idx:
                    # Destroy all widgets in the row
                    for widget in row_data["widgets"]:
                        widget.destroy()
                    # Remove the row from our list
                    self.rows.pop(i)
                    break

            # Reindex remaining rows
            for i, row_data in enumerate(self.rows, 1):
                row_data["row_index"] = i
                for j, widget in enumerate(row_data["widgets"]):
                    widget.grid(row=i, column=j)

            # Update amounts after deletion
            self.update_amounts()

            # Force update the UI
            self.update_idletasks()
            
        except Exception as e:
            logging.error(f"Error deleting row: {e}")
            messagebox.showerror("Error", "Failed to delete row. Please try again.")

    def clear_rows(self):
        """Clear all rows except one."""
        try:
            # Keep only the first row
            while len(self.rows) > 1:
                row_data = self.rows[-1]
                for widget in row_data["widgets"]:
                    widget.destroy()
                self.rows.pop()

            # Reset the first row
            first_row = self.rows[0]
            for widget in first_row["widgets"]:
                if isinstance(widget, (ctk.CTkEntry, ttk.Combobox)):
                    widget.delete(0, 'end')
                elif isinstance(widget, ctk.CTkLabel):
                    widget.configure(text="₹0.00")

            # Update amounts
            self.update_amounts()
            
            # Force update the UI
            self.update_idletasks()

        except Exception as e:
            logging.error(f"Error clearing rows: {e}")
            messagebox.showerror("Error", "Failed to clear rows. Please try again.")

    def update_amounts(self, event=None):
        """Update all amount calculations in real time (no debounce)."""
        self._do_update_amounts()

    def _do_update_amounts(self):
        """Actually perform the amount updates."""
        try:
            logging.debug("Updating amounts for all rows")
            total = 0.0
            mode = self.current_mode.get()

            # Calculate sum of row amounts
            for row_data in self.rows:
                widgets = row_data["widgets"]
                amount = 0.0 # Default amount
                try:
                    if mode == "Patti":
                        # Item [0], Pkt [1], Qty [2], Rate [3], Hamali [4], AmountLabel [5]
                        if len(widgets) > 4:
                            qty = validate_float(widgets[2].get())
                            rate = validate_float(widgets[3].get())
                            pkt = validate_float(widgets[1].get())
                            hamali_rate = validate_float(widgets[4].get())
                            # Calculate hamali based on number of packets
                            hamali_amount = pkt * hamali_rate
                            amount = (qty * rate) - hamali_amount  # Subtract hamali amount
                    elif mode == "Kata":
                        # Item [0], Net [1], Less% [2], Rate [3], HamaliRate [4], AmountLabel [5]
                         if len(widgets) > 4:
                            net = validate_float(widgets[1].get())
                            less = validate_float(widgets[2].get())
                            final_wt = net * (1 - less / 100.0) if less < 100 else 0.0
                            rate = validate_float(widgets[3].get())
                            hamali_rate = validate_float(widgets[4].get())
                            # Calculate packets based on net weight (e.g., if 60kg/packet)
                            packets = int(net / 60) if net > 0 else 0 
                            hamali_amount = packets * hamali_rate
                            amount = (final_wt * rate) - hamali_amount  # Subtract hamali amount
                    elif mode == "Barthe":
                        # Item [0], Pkt [1], Wt/Pkt [2], +/- Adj [3], Rate [4], Hamali/Pkt [5], AmountLabel [6]
                         if len(widgets) > 5:
                            pkt = validate_float(widgets[1].get())
                            wt_per_pkt = validate_float(widgets[2].get())
                            adj = validate_float(widgets[3].get()) # Adjustment weight
                            rate = validate_float(widgets[4].get()) # Rate per kg
                            hamali_rate = validate_float(widgets[5].get())
                            total_qty = (pkt * wt_per_pkt) + adj
                            # Calculate hamali based on number of packets
                            hamali_amount = pkt * hamali_rate
                            amount = (total_qty * rate) - hamali_amount  # Subtract hamali amount
                    
                    # Update the amount label for the current row
                    # The amount label is always the second-to-last widget (before delete button)
                    if len(widgets) > 0:
                        widgets[-2].configure(text=f"₹{amount:.2f}")
                    total += amount

                except IndexError:
                     logging.error(f"Index error calculating amount for row. Widgets: {len(widgets)}")
                except Exception as e:
                    logging.error(f"Error calculating amount: {e}")
                    if len(widgets) > 0:
                         widgets[-2].configure(text="₹Error") # Indicate error on the row

            # --- Add Kata Amount if applicable ---
            kata_amount = 0.0
            if mode == "Kata" and self.kata_amount_entry:
                try:
                    kata_amount = validate_float(self.kata_amount_entry.get())
                    # Add visual feedback for invalid input (optional)
                    if self.kata_amount_entry.get().strip() and kata_amount == 0 and self.kata_amount_entry.get() != '0':
                         self.kata_amount_entry.configure(fg_color="pink")
                    else:
                         # Reset color on valid input
                         self.kata_amount_entry.configure(fg_color=ctk.ThemeManager.theme["CTkEntry"]["fg_color"]) 
                except Exception as e:
                    logging.error(f"Error reading Kata amount: {e}")
                    # Maybe provide visual feedback on error
                    self.kata_amount_entry.configure(fg_color="pink")
            
            total -= kata_amount # Deduct validated Kata amount from total
            # --- End Add Kata Amount ---

            self.total_label.configure(text=f"Total Amount: ₹{total:.2f}")

        except Exception as e:
            error_msg = f"Error updating amounts: {str(e)}"
            logging.error(error_msg)
            self.total_label.configure(text="₹Error")

    def save_to_excel(self, show_popup=True, filename=None):
        try:
            # Get user's home directory
            home_dir = os.path.expanduser("~") 
            # Create the full path to the Documents folder
            documents_path = os.path.join(home_dir, "Documents")
            
            # Ensure the Documents directory exists, create if not
            os.makedirs(documents_path, exist_ok=True) 
            
            # Determine the filename
            if filename is not None:
                full_save_path = os.path.join(documents_path, filename)
            else:
                # Create the filename based on the current date
                date_str = datetime.now().strftime('%Y-%m-%d')
                base_filename = f"Invoice_{date_str}.xlsx"
                full_save_path = os.path.join(documents_path, base_filename)
            
            logging.info(f"Target save path: {full_save_path}")

            # Get Invoice Data
            customer = self.customer_entry.get().strip() or "Unknown Customer"
            mode = self.current_mode.get()
            
            headers = getattr(self, '_current_headers', []) 
            if not headers:
                if mode == "Patti":
                    headers = ["Item", "Packet", "Quantity", "Rate", "Hamali", "Amount"]
                elif mode == "Kata":
                    headers = ["Item", "Net Wt", "Less%", "Rate", "Hamali Rate", "Amount"]
                elif mode == "Barthe":
                    headers = ["Item", "Packet", "Weight", "+/-", "Rate", "Hamali", "Amount"]
                else: 
                    headers = ["Col1", "Col2", "Col3", "Col4", "Col5", "Col6", "Amount"]

            data_rows = []
            total_invoice_amount = 0.0
            kata_amount = 0.0
            if mode == "Kata" and self.kata_amount_entry:
                kata_amount = validate_float(self.kata_amount_entry.get())
            for row_data in self.rows:
                widgets = row_data["widgets"]
                row_values = []
                for w in widgets:
                    if isinstance(w, (ctk.CTkEntry, ttk.Combobox)):
                        row_values.append(w.get())
                    elif isinstance(w, ctk.CTkLabel):
                        row_values.append(w.cget("text").replace('₹', '').replace('Error', '0'))
                    else:
                        row_values.append("")

                if row_values and row_values[0].strip():
                    # For Kata mode, calculate total invoice amount
                    if mode == "Kata":
                        try:
                            total_invoice_amount += float(row_values[-2])
                        except Exception:
                            pass
                    data_rows.append(row_values)

            # For Kata mode, adjust the Amount column in each row to show the total
            if mode == "Kata" and data_rows:
                total_invoice_amount -= kata_amount
                for row in data_rows:
                    row[-2] = f"{total_invoice_amount:.2f}"

            if not data_rows:
                if show_popup:
                    messagebox.showwarning("No Data", "No data entered to save.")
                return

            # Excel Writing Logic with proper workbook handling
            try:
                if os.path.exists(full_save_path):
                    wb = load_workbook(full_save_path)
                else:
                    wb = Workbook()
                    
                # Check if mode sheet exists, create or get it
                if mode in wb.sheetnames:
                    ws = wb[mode]
                else:
                    if len(wb.sheetnames) > 0:
                        # If there are sheets but none match our mode, create new one
                        ws = wb.create_sheet(title=mode)
                    else:
                        # If it's a new workbook, rename the default sheet
                        ws = wb.active
                        ws.title = mode

                # Only write headers if the first row is empty or not matching
                expected_headers = ["Timestamp", "Customer"] + headers
                first_row = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
                if first_row != expected_headers:
                    ws.insert_rows(1)
                    for col, value in enumerate(expected_headers, 1):
                        ws.cell(row=1, column=col, value=value)

                # Write data
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                for row in data_rows:
                    ws.append([timestamp, customer] + row)

                # Save the workbook
                wb.save(full_save_path)
                logging.info(f"Successfully saved invoice data to {full_save_path} (Sheet: {mode})")
                if show_popup:
                    messagebox.showinfo("Saved", f"Invoice data saved to:\n{full_save_path}\n(Sheet: {mode})")

            except PermissionError:
                error_msg = f"Cannot save '{os.path.basename(full_save_path)}'.\nThe file might be open in Excel.\n\nLocation: {documents_path}"
                logging.error(error_msg)
                if show_popup:
                    messagebox.showerror("Permission Error", error_msg)
            except Exception as e:
                error_msg = f"Error saving Excel file to:\n{full_save_path}\n\nError: {str(e)}"
                logging.error(error_msg)
                if show_popup:
                    messagebox.showerror("Save Error", error_msg)

        except Exception as e:
            error_msg = f"Unexpected error during save operation: {str(e)}"
            logging.exception(error_msg)
            if show_popup:
                messagebox.showerror("Error", error_msg)

    def generate_print_content(self):
        """Generates the formatted string list for printing/preview."""
        lines = []
        customer = self.customer_entry.get().strip() or "N/A"
        mode = self.current_mode.get()
        
        max_width = 42  # Adjusted for 74mm thermal printer width

        # --- Header ---
        lines.extend([
            "G.V. Mahant Brothers".center(max_width),
            datetime.now().strftime("%d-%b-%Y %H:%M").center(max_width),
            "-" * max_width,
            f"Customer: {customer}".ljust(max_width),
            "-" * max_width,
        ])

        # --- Column Headers ---
        # Precise format strings for perfect alignment (42 chars total width)
        header_fmt = ""
        if mode == "Patti":
            # |Item   |Pkt |Qty  |Rate  |Ham |Amount |
            # |8chars|4chr|5chr |6chr  |4chr|8chars |
            header_fmt = "{:<8} {:>4} {:>5} {:>6} {:>4} {:>8}"
            headers = ["Item", "Pkt", "Qty", "Rate", "Ham", "Amt"]
            # Add separator line for visual alignment check
            lines.append("-" * max_width)
        elif mode == "Kata":
            # |Item   |Net |Less |Rate  |Ham |Amount |
            # |8chars|4chr|4chr |6chr  |4chr|8chars |
            header_fmt = "{:<8} {:>4} {:>4} {:>6} {:>4} {:>8}"
            headers = ["Item", "Net", "Less", "Rate", "Ham", "Amt"]
            # Add separator line for visual alignment check
            lines.append("-" * max_width)
        elif mode == "Barthe":
            # |Item  |Pkt |Wt  |+/-  |Rate |Ham |Amount |
            # |7chars|4chr|4chr|4chr |6chr |4chr|8chars |
            header_fmt = "{:<7} {:>4} {:>4} {:>4} {:>6} {:>4} {:>8}"
            headers = ["Item", "Pkt", "Wt", "+/-", "Rate", "Ham", "Amt"]
            # Add separator line for visual alignment check
            lines.append("-" * max_width)

        if header_fmt:
            lines.append(header_fmt.format(*headers))
        else:
            lines.append("Error: Mode not recognized for printing.")
        
        lines.append("-" * max_width)

        # --- Data Rows ---
        kata_amount_line = None
        total = 0.0
        for row_data in self.rows:
            widgets = row_data["widgets"]
            row_values = []
            for w in widgets:
                if isinstance(w, (ctk.CTkEntry, ttk.Combobox)):
                    row_values.append(w.get().strip())
                elif isinstance(w, ctk.CTkLabel):
                    text = w.cget("text").replace('₹', '').replace('Error', '0').strip()
                    row_values.append(text)
                else: 
                    row_values.append("")

            if not row_values or not row_values[0]:
                continue
            
            try:
                if mode == "Patti" and len(row_values) >= 6:
                    # Calculate hamali amount
                    pkt = float(row_values[1].replace('Error', '0').strip() or 0)
                    hamali_rate = float(row_values[4].replace('Error', '0').strip() or 0)
                    hamali_amount = int(pkt * hamali_rate)
                    amount_val = float(row_values[5].replace('₹', '').replace('Error', '0').strip() or 0)
                    amount_str = f"{amount_val:.2f}"
                    lines.append(header_fmt.format(
                        row_values[0][:8],        # Item name (left-aligned, 8 chars)
                        row_values[1][:4],        # Pkt (right-aligned, 4 chars)
                        row_values[2][:5],        # Qty (right-aligned, 5 chars)
                        row_values[3][:6],        # Rate (right-aligned, 6 chars)
                        str(hamali_amount)[:4],   # Hamali (right-aligned, 4 chars)
                        amount_str[:8]            # Amount (right-aligned, 8 chars)
                    ))
                    total += amount_val
                elif mode == "Kata" and len(row_values) >= 6:
                    # Calculate hamali amount
                    net = float(row_values[1].replace('Error', '0').strip() or 0)
                    packets = int(net / 60) if net > 0 else 0
                    hamali_rate = float(row_values[4].replace('Error', '0').strip() or 0)
                    hamali_amount = int(packets * hamali_rate)
                    amount_val = float(row_values[5].replace('₹', '').replace('Error', '0').strip() or 0)
                    amount_str = f"{amount_val:.2f}"
                    lines.append(header_fmt.format(
                        row_values[0][:8],        # Item name (left-aligned, 8 chars)
                        row_values[1][:4],        # Net (right-aligned, 4 chars)
                        row_values[2][:4],        # Less (right-aligned, 4 chars)
                        row_values[3][:6],        # Rate (right-aligned, 6 chars)
                        str(hamali_amount)[:4],   # Hamali (right-aligned, 4 chars)
                        amount_str[:8]            # Amount (right-aligned, 8 chars)
                    ))
                    total += amount_val
                elif mode == "Barthe" and len(row_values) >= 7:
                    # Calculate hamali amount
                    pkt = float(row_values[1].replace('Error', '0').strip() or 0)
                    hamali_rate = float(row_values[5].replace('Error', '0').strip() or 0)
                    hamali_amount = int(pkt * hamali_rate)
                    amount_val = float(row_values[6].replace('₹', '').replace('Error', '0').strip() or 0)
                    amount_str = f"{amount_val:.2f}"
                    lines.append(header_fmt.format(
                        row_values[0][:7],        # Item name (left-aligned, 7 chars)
                        row_values[1][:4],        # Pkt (right-aligned, 4 chars)
                        row_values[2][:4],        # Weight (right-aligned, 4 chars)
                        row_values[3][:4],        # +/- (right-aligned, 4 chars)
                        row_values[4][:6],        # Rate (right-aligned, 6 chars)
                        str(hamali_amount)[:4],   # Hamali (right-aligned, 4 chars)
                        amount_str[:8]            # Amount (right-aligned, 8 chars)
                    ))
                    total += amount_val
            except Exception as fmt_e:
                lines.append(f"Fmt Error: {fmt_e}")

        # --- Add Kata Amount if applicable ---
        if mode == "Kata" and self.kata_amount_entry:
            kata_val_str = self.kata_amount_entry.get().strip()
            kata_amount = validate_float(kata_val_str)
            # Place Kata Amount at the far left, below the last data row
            kata_amount_line = f"Kata Amount: {kata_amount:.2f}"
            lines.append(kata_amount_line.ljust(max_width))
            total -= kata_amount  # Deduct kata amount from total

        # --- Footer ---
        lines.extend([
            "-" * max_width,
            f"Total Amount: ₹{total:.2f}".center(max_width),
            "-" * max_width,
            "\n",  # Line feed
            "\n",  # Line feed
            "\n"   # Line feed
        ])

        # Add printer cut command (ESC/POS standard)
        lines.append(chr(27) + chr(105)) # Full cut

        return lines

    def save_for_print(self):
        """Prints the generated content to the default printer."""
        try:
            printer_name = win32print.GetDefaultPrinter()
            logging.info(f"Attempting to print to default printer: {printer_name}")
            
            lines = self.generate_print_content()
            print_content = "\n".join(lines)
            
            # First try UTF-8 encoding for Kannada text
            try:
                print_bytes = print_content.encode('utf-8')
            except UnicodeEncodeError:
                # If UTF-8 fails, try UTF-16
                try:
                    print_bytes = print_content.encode('utf-16')
                except UnicodeEncodeError:
                    # If both fail, fall back to cp437 with replacement
                    print_bytes = print_content.encode('cp437', errors='replace')

            # Use win32print for direct RAW printing
            hPrinter = win32print.OpenPrinter(printer_name)
            try:
                # Job name "Invoice", Datatype "RAW"
                hJob = win32print.StartDocPrinter(hPrinter, 1, ("Invoice", None, "RAW"))
                try:
                    win32print.StartPagePrinter(hPrinter)
                    win32print.WritePrinter(hPrinter, print_bytes)
                    win32print.EndPagePrinter(hPrinter)
                finally:
                    win32print.EndDocPrinter(hPrinter)
            finally:
                win32print.ClosePrinter(hPrinter)

            logging.info("Invoice successfully sent to printer.")
            messagebox.showinfo("Success", "Invoice sent to printer!")

        except Exception as e:
            error_msg = f"Error printing invoice: {str(e)}"
            logging.error(error_msg)
            logging.error("Printer encoding error - trying to print Kannada text")
            messagebox.showerror("Print Error", f"Could not print to {printer_name}.\nCheck if your printer supports Kannada text.\n\nError: {e}")


    def show_print_preview(self):
        """Shows a Toplevel window with a preview of the print output."""
        try:
            # Auto-save before showing preview
            self.save_to_excel(show_popup=False)
            
            preview = ctk.CTkToplevel(self)
            preview.title("Print Preview")
            preview.geometry("450x600") # Slightly wider for better view
            preview.transient(self) # Keep preview on top of main window
            preview.grab_set()  # Make the window modal

            # Center the preview window relative to the main app
            app_x = self.winfo_x()
            app_y = self.winfo_y()
            app_w = self.winfo_width()
            app_h = self.winfo_height()
            pre_w = 450
            pre_h = 600
            x = app_x + (app_w - pre_w) // 2
            y = app_y + (app_h - pre_h) // 2
            preview.geometry(f"{pre_w}x{pre_h}+{x}+{y}")

            # --- Preview Content Area ---
            # Use a CTkTextbox within a ScrollableFrame for better handling
            scroll_frame = ctk.CTkScrollableFrame(preview, label_text="Preview")
            scroll_frame.pack(fill="both", expand=True, padx=10, pady=(10, 0))

            preview_text = ctk.CTkTextbox(
                scroll_frame, 
                font=("Courier New", 10), # Monospace font for alignment
                wrap="none" # Prevent wrapping to see true line breaks
            )
            preview_text.pack(fill="both", expand=True)
            
            # Generate print content and display it
            lines = self.generate_print_content()
            # Join lines, but remove the final cut command for preview
            preview_content = "\n".join(lines[:-1]) if lines else "" 
            
            preview_text.insert("1.0", preview_content)
            preview_text.configure(state="disabled")  # Make read-only

            # --- Buttons Frame ---
            button_frame = ctk.CTkFrame(preview, fg_color="transparent")
            button_frame.pack(fill="x", padx=10, pady=10)
            
            # Center buttons using grid
            button_frame.grid_columnconfigure(0, weight=1)
            button_frame.grid_columnconfigure(1, weight=1)

            ctk.CTkButton(
                button_frame,
                text="Print",
                # Lambda calls destroy first, then the print function
                command=lambda: [preview.destroy(), self.save_for_print()], 
                width=120
            ).grid(row=0, column=0, padx=5, pady=5, sticky="ew")

            ctk.CTkButton(
                button_frame,
                text="Close",
                command=preview.destroy,
                width=120
            ).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
            
            preview.after(100, preview.lift) # Ensure it's raised above main window

        except Exception as e:
            error_msg = f"Error generating print preview: {str(e)}"
            logging.error(error_msg)
            messagebox.showerror("Preview Error", error_msg)
            if 'preview' in locals() and preview.winfo_exists():
                 preview.destroy() # Close broken preview window

    def update_datetime(self):
        """Update the date/time label with current time."""
        # Update every 5 seconds instead of every second
        current_time = datetime.now().strftime("%I:%M:%S %p")
        current_date = datetime.now().strftime("%A, %d %B %Y")
        if not hasattr(self, '_last_date') or self._last_date != current_date:
            self._last_date = current_date
            self.date_label.configure(text=f"{current_date} {current_time}")
        else:
            # Only update the time portion
            self.date_label.configure(text=f"{self._last_date} {current_time}")
        self.after(5000, self.update_datetime)  # Update every 5 seconds

    def set_mode(self, mode):
        """Set the current mode and update button colors."""
        self.current_mode.set(mode)
        
        # Update button colors for all modes
        for btn_mode, btn in self.mode_buttons.items():
            if btn_mode == mode:
                btn.configure(
                    fg_color="#547792",
                    text_color="white",
                    border_color="#547792",
                    hover_color="#63889e"
                )
            else:
                btn.configure(
                    fg_color="#e0e0e0",
                    text_color="#212121",
                    border_color="#d0d0d0",
                    hover_color="#d0d0d0"
                )
        
        # Update table headers and other UI elements
        self.switch_mode()

    def only_numeric_input(self, P):
        # Allow empty string, integer, or float
        if P == "" or (P.replace('.', '', 1).isdigit() and P.count('.') <= 1):
            return True
        return False

    def debounce_update_amounts(self, event=None):
        if hasattr(self, '_debounce_after_id') and self._debounce_after_id:
            self.after_cancel(self._debounce_after_id)
        self._debounce_after_id = self.after(300, self._do_update_amounts)  # 300ms debounce

    def select_all_on_focus(self, event):
        event.widget.select_range(0, 'end')
        event.widget.icursor('end')

    def save_to_excel_async(self):
        threading.Thread(target=self.save_to_excel, daemon=True).start()

    def save_for_print_async(self):
        threading.Thread(target=self.save_for_print, daemon=True).start()

    def auto_save(self):
        # Save to a special autosave file
        self.save_to_excel(filename='autosave_invoice.xlsx', show_popup=False)

    def check_autosave_on_start(self):
        if os.path.exists('autosave_invoice.xlsx'):
            if messagebox.askyesno("Recover?", "Recover unsaved invoice from last session?"):
                self.load_invoice('autosave_invoice.xlsx')

    def load_invoice(self, filename):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename)
            ws = wb.active
            # Clear current rows
            self.clear_rows()
            # Read the first data row (skip header)
            first_data_row = None
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if i == 0:
                    # Set customer name from first row
                    self.customer_entry.delete(0, 'end')
                    self.customer_entry.insert(0, row[1])
                # Add a row and fill in values
                self.add_row()
                last_row = self.rows[-1]['widgets']
                # Fill in item and entry fields (skip timestamp, customer, and amount)
                for j, value in enumerate(row[2:-1]):
                    if hasattr(last_row[j], 'delete'):
                        last_row[j].delete(0, 'end')
                        last_row[j].insert(0, str(value) if value is not None else "")
            self.update_amounts()
            import os
            os.remove(filename)  # Remove autosave after recovery
            messagebox.showinfo('Recovered', 'Invoice data recovered from autosave.')
        except Exception as e:
            messagebox.showerror('Error', f'Could not load invoice: {e}')

    def schedule_auto_save(self):
        self.auto_save()
        self.after(900000, self.schedule_auto_save)  # 15 minutes (900,000 ms)

if __name__ == "__main__":
    app = InvoiceApp()
    app.mainloop()